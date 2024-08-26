# ROCReaderGenerator.py
#
# Copyright (C) 2024-2030 Yun Liu
# University of Chicago
#
# LaTeX ROC Reader Generator
#

import sys
import os
import re
import pandas as pd
import string
import json
from openpyxl import load_workbook
from ROCAveGenerator import LaTeXROCAveGenerator

class LaTeXROCReaderGenerator(LaTeXROCAveGenerator):
    def __init__(self):
        super().__init__()
        self.reader_files = []
        self.reader_groups = []
        self.reader_index_map = {}
        
        self.sheet_names = set()
        self.reader_colors = []  
        self.reader_marks = []
        self.default_styles = ['square*', '*', 'dot', 'triangle*', 'x', 'pentagon*']

        self.reader_data_commands = ""
        self.reader_color_definitions = ""
        self.reader_plot_commands = ""
        self.reader_plot_frame_commands = ""
        self.plot_reader_group_commands = ""
        self.plot_reader_fig_command = ""
        self.make_figure_command = ""

    def clean_name(self, name):
        """
        Cleans the group or sheet name by removing non-alphanumeric characters.
        """
        return re.sub(r'[^a-zA-Z0-9]', '', name)

    def generate_reader_index_map(self):
        """
        Generates an index map for the reader ROC groups and their respective data sheets.
        """
        index_map = {}
        for type_name, methods in self.reader_groups:
            for method_name, file, sheets in methods:
                for sheet_name in sheets:
                    cmd_index = f"{type_name}_{method_name}_{sheet_name}"
                    letter_index = f"{method_name}{sheet_name}{type_name}"
                    index_map[cmd_index] = letter_index
        return index_map

    def parse_reader_files(self, reader_files, type_names=None, group_names=None, engine='openpyxl'):
        """
        Parses multiple Excel files to extract reader ROC data and organize it into groups.
        """
        groups = []
        type_names = type_names or list(self.generate_names(len(reader_files)))
        
        if group_names is None:
            group_names = [list(self.generate_names(len(type_files))) for type_files in reader_files]
   
        for type_files, type_name, methods in zip(reader_files, type_names, group_names):
            if not all(os.path.exists(file) and file.endswith('.xlsx') for file in type_files):
                raise ValueError("One or more files do not exist or are not XLSX files.")
            
            method_sheets = []
            for file, method_name in zip(type_files, methods):
                sheets = pd.ExcelFile(file, engine=engine).sheet_names
                self.sheet_names.update(sheets)
                method_sheets.append((method_name, file, sheets))
            groups.append((type_name, method_sheets))
        
        self.reader_files = reader_files
        self.reader_groups = groups
        self.reader_index_map = self.generate_reader_index_map()

        max_files_per_type = max(len(methods) for _, methods in self.reader_groups)
        if not self.reader_colors:
            self.reader_colors = self.default_colors[:max_files_per_type]
        self.reader_marks = [self.default_styles[:max_files_per_type]] * len(type_names)

        self.reader_data_commands = self.generate_reader_data_commands()
        self.reader_color_definitions = self.generate_reader_color_definitions()
        self.reader_plot_commands = self.generate_reader_plot_commands()
        self.reader_plot_frame_commands = self.generate_reader_plot_frame_commands()
        self.make_figure_command = self.generate_make_figure_command()
        self.plot_fig_command = self.generate_plot_fig_command()
        self.plot_reader_group_commands = self.generate_reader_group_commands()
        self.header_footer = self.generate_header_footer()

    def generate_reader_data_commands(self):
        """
        Generates LaTeX commands for reader ROC data from the parsed Excel files.
        """
        data_commands = "% Reader ROC curve and AUC data:\n"
        for type_name, methods in self.reader_groups:
            for method_name, file, sheets in methods:
                for sheet_name in sheets:
                    wb = load_workbook(file)
                    ws = wb[sheet_name]
                    auc_value = ws['B1'].value
                    roc_data = [(row[0], row[1]) for row in ws.iter_rows(min_row=3, values_only=True)]
                    cmd_index = f"{type_name}_{method_name}_{sheet_name}"
                    letter_index = self.reader_index_map[cmd_index]
                    data_commands += f"\\newcommand{{\\{letter_index}AUC}}[0]{{{auc_value}}}\n"
                    data_commands += f"\\newcommand{{\\{letter_index}Data}}[0]{{\n  " + "\n  ".join([f"({x[0]},{x[1]})" for x in roc_data]) + "\n}\n\n"
        return data_commands
    
    def generate_reader_color_definitions(self):
        """
        Generates LaTeX commands to define colors for reader ROC curves.
        """
        color_definitions = "% Define ROC curve colors:\n"
        for type_index, (type_name, methods) in enumerate(self.reader_groups):
            for method_index, (method_name, file, sheets) in enumerate(methods):
                color_index = method_index % len(self.reader_colors)
                color = self.reader_colors[color_index]
                color_definitions += f"\\definecolor{{COLORo{type_name}{method_name}}}{{rgb}}{color}\n"
        color_definitions += "\n"
        return color_definitions

    def generate_reader_plot_commands(self):
        """
        Generates LaTeX commands for plotting individual reader ROC curves.
        """
        plot_style = "% Command for plotting reader ROC curves:\n"
        is_first_plot = 1
        for type_index, (type_name, methods) in enumerate(self.reader_groups):
            for method_index, (method_name, file, sheets) in enumerate(methods):
                mark_type = self.reader_marks[type_index][method_index]
                for sheet_index, sheet_name in enumerate(sheets):
                    cmd_index = f"{type_name}_{method_name}_{sheet_name}"
                    letter_index = self.reader_index_map[cmd_index]

                    wb = load_workbook(file)
                    ws = wb[sheet_name]
                    row_count = ws.max_row
                    break
                
                if is_first_plot:
                    plot_style += f"\\newcommand{{\\{type_name}{method_name}}}[3]{{\n"
                    plot_style += f"\\addlegendimage{{empty legend}}\n"
                    plot_style += f"\\addlegendentry{{Reader #3}}\n"
                    is_first_plot -= 1           
                else:
                    plot_style += f"\\newcommand{{\\{type_name}{method_name}}}[2]{{\n"  
                    
                if row_count <= 50:         
                    plot_style += f"\\addplot[\n  color=COLORo{type_name}{method_name},\n  only marks,\n  mark={mark_type},\n  on layer={{axis foreground}},\n] coordinates {{#1}};\n"
                else:
                    plot_style += f"\\addplot[\n  color=COLORo{type_name}{method_name},\n  mark=dot,\n  on layer={{axis foreground}},\n] coordinates {{#1}};\n"
                    self.reader_marks[type_index][method_index] = 'dot'
                    
                plot_style += f"\\addlegendentry{{#2}}\n}}\n\n"
        return plot_style
    
    def generate_reader_plot_frame_commands(self):
        """
        Generates LaTeX commands for plotting grouped reader ROC curves.
        """
        plot_frame_commands = "% Commands for plotting ROC curves for every reader:\n"

        for sheet_name in sorted(self.sheet_names):
            clean_sheet_name = self.clean_name(sheet_name)
            plot_frame_commands += f"\\newcommand{{\\R{clean_sheet_name}}}{{\n"
            first_entry = True
            for type_name, methods in self.reader_groups:
                for method_name, file, sheets in methods:
                    if sheet_name in sheets:
                        cmd_index = f"{type_name}_{method_name}_{sheet_name}"
                        letter_index = self.reader_index_map[cmd_index]
                        plot_frame_commands += f"  \\{type_name}{method_name}{{\\{letter_index}Data}}{{\\{letter_index}AUC}}"
                        if first_entry:  
                            plot_frame_commands += f"{{{sheet_name}}}"
                            first_entry = False
                        plot_frame_commands += "\n"
            plot_frame_commands += "}\n\n"
        
        return plot_frame_commands


#     def generate_make_reader_figure_command(self):
#         tick_style = ", ".join([f"{k}={v}" for k, v in self.plot_format["tick_style"].items()])
#         legend_style = ", ".join([f"{k}={v}" for k, v in self.plot_format["legend_style"].items()])
#         plot_format_combined = self.plot_format.copy()
#         plot_format_combined.update({"tick_style": tick_style, "legend_style": legend_style})
#         return """
# % Command for plotting a figure of a group of ROC curves:
# \\newcommand{{\\ROCPlotsOnePage}}[1]{{
# \\begin{{center}}
# \\begin{{tikzpicture}}
# \\begin{{groupplot}}[
#   group style={{group size={horizontal_size} by {vertical_size},
#   horizontal sep={horizontal_sep}, vertical sep={vertical_sep}}},
#   width={width},
#   height={height},
#   label style={{font={{\\fontsize{{{label_font_size_1}}}{{{label_font_size_2}}}\\selectfont}}}},
#   domain={domain},
#   restrict y to domain={restrict_y_domain},
#   samples={samples},
#   minor tick num={minor_tick_num},
#   xmin={xmin}, xmax={xmax},
#   ymin={ymin}, ymax={ymax},
#   xticklabels={{,,}},
#   yticklabels={{,,}},
#   tick label style={{font={{\\fontsize{{{tick_label_font_size_1}}}{{{tick_label_font_size_2}}}\\selectfont}}}},
#   tick style={{{tick_style}}},
#   legend style={{{legend_style}}},
#   set layers=standard,
# ] #1
# \\end{{groupplot}}
# \\end{{tikzpicture}}
# \\end{{center}}
# }}

# """.format(**plot_format_combined)

    def generate_plot_fig_command(self):
        """
        Generates LaTeX commands for plotting figures in a 4x3 grid.
        """
        return r"""
% Commands for plotting 4 rows of 3 ROC curves:
\newcommand{\RowA}[3]{
\nextgroupplot[
  ylabel = {{\textbf{{TPF}}}},
  yticklabels={{, , 0.2, 0.4, 0.6, 0.8, 1.0}},
] #1
\nextgroupplot[] #2
\nextgroupplot[] #3
}

\newcommand{\RowB}[3]{
\nextgroupplot[
  ylabel = {{\textbf{{TPF}}}},
  yticklabels={{, , 0.2, 0.4, 0.6, 0.8, 1.0}},
] #1
\nextgroupplot[] #2
\nextgroupplot[] #3
}

\newcommand{\RowC}[3]{
\nextgroupplot[
  ylabel = {{\textbf{{TPF}}}},
  yticklabels={{, , 0.2, 0.4, 0.6, 0.8, 1.0}},
] #1
\nextgroupplot[] #2
\nextgroupplot[] #3
}

\newcommand{\RowD}[3]{
\nextgroupplot[
  xlabel = {\textbf{{FPF}}},
  ylabel = {\textbf{{TPF}}},
  xticklabels={, 0.0, 0.2, 0.4, 0.6, 0.8, 1.0},
  yticklabels={, 0.0, 0.2, 0.4, 0.6, 0.8, 1.0},
] #1
\nextgroupplot[
  xlabel = {\textbf{{FPF}}},
  xticklabels={, , 0.2, 0.4, 0.6, 0.8, 1.0},
] #2
\nextgroupplot[
  xlabel = {\textbf{{FPF}}},
  xticklabels={, , 0.2, 0.4, 0.6, 0.8, 1.0},
] #3
}

"""

    def generate_reader_group_commands(self):
        """
        Generates LaTeX commands for plotting ROC curves for each group of readers.
        """
        group_commands = ""
        sheet_names = sorted(self.sheet_names)
        num_sheets = len(sheet_names)
        num_pages = (num_sheets + 11) // 12 

        for page in range(num_pages):
            start = page * 12
            end = min(start + 12, num_sheets)
            group_commands += f"% Command for plotting page {page + 1} of ROC curves for 24 readers:\n"
            group_commands += f"\\newcommand{{\\ROCplotPage{chr(65 + page)}}}[0]{{\n"  
            for row in range(4):
                row_letter = chr(ord('A') + row)
                group_commands += f"  \\Row{row_letter}{{"
                for col in range(3):
                    idx = start + row * 3 + col
                    if idx < num_sheets:
                        clean_sheet_name = self.clean_name(sheet_names[idx])
                        group_commands += f"\\R{clean_sheet_name}"
                    else:
                        group_commands += ""
                    if col < 2:
                        group_commands += "}{"
                group_commands += "}\n"
            group_commands += "}\n\n"
        return group_commands
    
    def generate_document_body_commands(self):
        """
        Generates LaTeX commands for the body of the document.
        """
        body_commands = r""
        sheet_names = sorted(self.sheet_names)
        num_sheets = len(sheet_names)
        num_pages = (num_sheets + 11) // 12 
        for page in range(num_pages):
            body_commands += f"\\MakeAfigure{{\\ROCplotPage{chr(65 + page)}}}\n"
        return body_commands

    def generate_document_body(self):
        """
        Generates the complete LaTeX document body.
        """
        body = r"\begin{document}"
        body += self.generate_document_body_commands()
        body += "\n"
        body += r"\end{document}"
        return body

    def generate_latex_document(self):
        """
        Generates the complete LaTeX document.
        """
        latex_document = (
            self.generate_document_header() +
            self.reader_data_commands +
            self.reader_color_definitions +
            self.reader_plot_commands +
            self.reader_plot_frame_commands +
            self.make_figure_command +
            self.plot_fig_command +
            self.plot_reader_group_commands +
            self.header_footer +
            self.generate_document_body()
        )
        return latex_document

    def generate_latex_image(self):
        """
        Generates the LaTeX document for standalone images.
        """
        latex_document = (
            self.generate_image_header() +
            self.reader_data_commands +
            self.reader_color_definitions +
            self.reader_plot_commands +
            self.reader_plot_frame_commands +
            self.make_figure_command +
            self.plot_fig_command +
            self.plot_reader_group_commands +
            self.generate_document_body()
        )
        return latex_document

    def export_reader_settings(self):
        """
        Exports the current settings of page_format, plot_format, and reader_colors.
        """
        settings = {
            'page_format': self.page_format,
            'plot_format': self.plot_format,
            'reader_colors': self.reader_colors
        }
        return settings
    
    def import_reader_settings(self, settings):
        """
        Imports settings from a JSON file to page_format, plot_format, and reader_colors.
        """
        if isinstance(settings, str):
            with open(settings, 'r') as file:
                settings = json.load(file)
        self.page_format = settings['page_format']
        self.plot_format = settings['plot_format']
        self.reader_colors = settings['reader_colors']
            
        self.document_header = self.generate_document_header()
        self.make_figure_command = self.generate_make_figure_command()    
        self.reader_color_definitions = self.generate_reader_color_definitions()

if __name__ == "__main__":
    try:
        reader_files = [
            ["Data/QR_t_Data.xlsx", "Data/MR_t_Data.xlsx"],
            ["Data/QR_p_Data.xlsx", "Data/MR_p_Data.xlsx"]
        ]

        type_names = ["t", "p"]
        group_names = [["QR", "MR"], ["QR", "MR"]]
        
        # reader_files = [
        #     ["Data/QR_t_Data.xlsx", "Data/MR_t_Data.xlsx"],
        #     ["Data/QR_p_Data.xlsx", "Data/MR_p_Data.xlsx"],
        #     ["Data/QR_p_Data.xlsx", "Data/MR_p_Data.xlsx"]
        # ]

        # type_names = ["t", "p", "x"]
        # group_names = [["QR", "MR"], ["QR", "MR"], ["QR", "MR"]]
        
        generator = LaTeXROCReaderGenerator()

        generator.parse_reader_files(reader_files, type_names, group_names)
        # generator.set_header_info(author="New Author")
        # generator.set_page_format(margin=".75in", top_margin=".8in")
        # generator.set_plot_format(width="3in", height="3in", tick_style={"draw": "red"}, legend_style={"anchor": "west"})

        generator.set_plot_format(legend_style={"at": "{(0.4,0.3)}"})
        generator.set_plot_format(x_ticklabels= "{,,}", y_ticklabels= "{,,}")
        generator.set_header_info(name="ROC Reader Analysis")

        # Generate full document
        latex_document = generator.generate_latex_document()
        output_dir = 'Output'
        os.makedirs(output_dir, exist_ok=True)

        doc_file_path = os.path.join(output_dir, 'ROC_reader_analysis.tex')
        with open(doc_file_path, 'w') as f:
            f.write(latex_document)

        image = generator.generate_latex_image()
        image_file_path = os.path.join(output_dir, 'ROC_reader_image.tex')
        with open(image_file_path, 'w') as f:
            f.write(image)

    except ValueError as e:
        print(e)
        sys.exit(1)
    except Exception as e:
        print(f"Unexpected error: {e}")
        sys.exit(1)