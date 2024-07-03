import sys
import os
import string
from openpyxl import load_workbook
from ROCGenerator import LaTeXROCGenerator 

class LaTeXROCReaderAveGenerator(LaTeXROCGenerator):
    def __init__(self):
        super().__init__()
        self.reader_files = []
        self.reader_data_commands = ""
        self.reader_plot_commands = ""
        self.readerave_plot_frame_commands = ""
        self.plot_readerave_group_commands = ""
        self.plot_reader_fig_command = ""

    def generate_ave_index_map(self):
        index_map = {}
        counter = 0
        for group_index, group in enumerate(self.avef_groups):
            for file_index, _ in enumerate(group):
                cmd_index = f"{group_index}_{file_index}"
                letter_index = string.ascii_uppercase[counter % len(string.ascii_lowercase)]
                index_map[cmd_index] = letter_index
                counter += 1
        return index_map
    
    def parse_reader_files(self, reader_files):
        files = []
        for file in reader_files:
            if not os.path.exists(file) or not file.endswith('.xlsx'):
                raise ValueError(f"Error: File '{file}' does not exist或 is not an Excel file.")
            files.append(file)
        self.reader_files = files
        self.reader_data_commands = self.generate_reader_data_commands()
        self.reader_plot_commands = self.generate_reader_plot_commands()
        self.readerave_plot_frame_commands = self.generate_readerave_plot_frame_commands()
        self.plot_readerave_group_commands = self.generate_readerave_group_commands()

    def generate_reader_data_commands(self):
        data_commands = "% Reader ROC curve and AUC data:\n"
        for group_index, excel_file in enumerate(self.reader_files):
            wb = load_workbook(excel_file)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                auc_value = ws['B1'].value
                roc_data = [(row[0], row[1]) for row in ws.iter_rows(min_row=3, values_only=True)]
                letter_index = f"{string.ascii_uppercase[group_index]}{sheet_name[0]}"
                data_commands += f"\\newcommand{{\\{letter_index}AUC}}[0]{{{auc_value}}}\n"
                data_commands += f"\\newcommand{{\\{letter_index}Data}}[0]{{\n  " + "\n  ".join([f"({x[0]},{x[1]})" for x in roc_data]) + "\n}\n\n"
        return data_commands

    def generate_ave_plot_commands(self):
        plot_commands = "% Command for plotting ROC curves in one plot:\n"
        for group_index, group in enumerate(self.avef_groups):
            for file_index, _ in enumerate(group):
                cmd_index = f"{group_index}_{file_index}"
                letter_index = self.avef_index_map[cmd_index]
                plot_commands += f"\\newcommand{{\\DrawLINE{letter_index}Ave}}[2]{{\n"
                plot_commands += f"\\addplot[\n  color=COLORo{file_index},\n  mark=dot,\n  line width=2pt,\n  on layer={{axis foreground}},\n] coordinates {{#1}};\n"
                plot_commands += f"\\addlegendentry{{#2}}\n}}\n\n"
        return plot_commands

    def generate_reader_plot_commands(self):
        reader_plot_commands = "% Command for plotting a reader ROC curve:\n"
        for group_index, excel_file in enumerate(self.reader_files):
            letter_index = f"{string.ascii_uppercase[group_index]}"
            reader_plot_commands += f"\\newcommand{{\\DrawLINE{letter_index}Reader}}[1]{{\n"
            reader_plot_commands += f"\\addplot[\n  color=COLORo{(group_index + 1) % self.num_ave_colors},\n  mark=dot,\n  line width=.5pt,\n  on layer={{axis foreground}},\n] coordinates {{#1}};\n}}\n\n"
        return reader_plot_commands
    
    def generate_readerave_plot_frame_commands(self):
        plot_frame_commands = "% Command for plotting ROC curves in one plot:\n"
        for group_index, excel_file in enumerate(self.reader_files):
            group_letter = string.ascii_uppercase[group_index]
            plot_frame_commands += f"\\newcommand{{\\PlotFRAMEoReadernAve{group_letter}}}{{\n"
            
            wb = load_workbook(excel_file)
            for sheet_name in wb.sheetnames:
                letter_index = f"{string.ascii_uppercase[group_index]}{sheet_name[0]}"
                plot_frame_commands += f"  \\DrawLINE{group_letter}Reader{{\\{letter_index}Data}}\n"
            plot_frame_commands += f"  \\DrawLINE{group_letter}Ave{{\\DATAoROC{group_letter}}} {{Average {group_letter}}}\n"
            plot_frame_commands += "}\n\n"
        return plot_frame_commands

    def generate_readerave_group_commands(self):
        group_commands = "%Command for plotting group figures"
        for group_index, _ in enumerate(self.reader_files):
            group_letter = string.ascii_uppercase[group_index]
            group_commands += f"""
\\newcommand{{\\PlotFIGoReadernAve{group_letter}}}[0]{{
  \\PlotFIG{{\\PlotFRAMEoReadernAve{group_letter}}}
}}
"""
        return group_commands


    def generate_document_body(self):
        body = "% Document body:\n"
        body += "\\begin{document}\n"
        for group_index, _ in enumerate(self.reader_files):
            group_letter = string.ascii_uppercase[group_index]
            body += f"\\MakeAfigure{{\\PlotFIGoReadernAve{group_letter}}}\n"
        body += "\\end{document}"
        return body

    def generate_latex_document(self):
        latex_document = (
            self.generate_document_header() +
            self.ave_data_commands +
            self.reader_data_commands +
            self.header_footer +
            self.ave_color_definitions +
            self.ave_plot_commands +
            self.reader_plot_commands +
            self.readerave_plot_frame_commands +
            self.plot_readerave_group_commands +
            self.plot_fig_command +
            self.make_ave_figure_command +

            self.generate_document_body()
        )
        return latex_document

if __name__ == "__main__":
    try:
        group_files = ["Data/NP.xlsx", "Data/PBN.xlsx"]
        reader_files = [
            ["Data/QR_t_Data.xlsx", "Data/MR_t_Data.xlsx"],
            ["Data/QR_p_Data.xlsx", "Data/MR_p_Data.xlsx"]
        ]        
        
        generator = LaTeXROCReaderAveGenerator()
        
        generator.parse_ave_group_files(group_files)
        generator.parse_reader_files(reader_files)

        generator.set_ave_colors((1.0, 0.0, 0.0), (0.0, 0.0, 1.0)) 
        generator.set_header_info(author="New Author")

        generator.set_page_format(margin=".75in", top_margin=".8in")
        generator.set_plot_format(width="3in", height="3in", tick_style={"draw": "red"}, legend_style={"anchor": "west"})

        latex_document = generator.generate_latex_document()
        with open('Output/reader_ave_output.tex', 'w') as f:
            f.write(latex_document)

    except ValueError as e:
        print(e)
        sys.exit(1)
    except Exception as e:
        print(f"Unexpected error: {e}")
        sys.exit(1)